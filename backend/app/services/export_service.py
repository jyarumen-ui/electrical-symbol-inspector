import io
from datetime import date, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from ..schemas.estimation import EstimationSummary


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")

_thin = Side(style="thin")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _jp(text: str) -> str:
    return text


def build_excel(summary: EstimationSummary, company_name: str, site_name: str) -> bytes:
    wb = Workbook()

    # ── Sheet 1: 見積明細 ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "見積明細"
    _write_estimation_sheet(ws1, summary, company_name, site_name)

    # ── Sheet 2: 資材発注リスト ────────────────────────────────────────
    ws2 = wb.create_sheet("資材発注リスト")
    _write_order_sheet(ws2, summary)

    # ── Sheet 3: 未確定・要確認 ───────────────────────────────────────
    ws3 = wb.create_sheet("未確定・要確認")
    _write_pending_sheet(ws3, summary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_estimation_sheet(ws, summary: EstimationSummary, company_name: str, site_name: str):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 20

    # Header block
    ws.merge_cells("A1:I1")
    ws["A1"] = "見　積　書"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "会社名:"
    ws["B2"] = company_name
    ws["F2"] = "現場名:"
    ws["G2"] = site_name
    ws["F3"] = "作成日:"
    ws["G3"] = date.today().strftime("%Y年%m月%d日")
    ws["F4"] = "有効期限:"
    ws["G4"] = (date.today() + timedelta(days=30)).strftime("%Y年%m月%d日")

    # Column headers row 6
    headers = ["記号コード", "品名", "メーカー", "型番", "数量", "単位", "単価", "金額", "備考"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border

    row = 7
    all_items = summary.confirmed_items + summary.pending_items
    for item in all_items:
        is_pending = item.status == "PENDING"
        fill = PENDING_FILL if is_pending else None
        values = [
            item.symbol_code,
            item.item_name + (" [要確認]" if is_pending else ""),
            item.maker or "",
            item.model_number or "",
            float(item.quantity),
            item.unit,
            float(item.unit_price),
            float(item.amount),
            item.note or "",
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _border
            if fill:
                c.fill = fill
            if col in (7, 8):
                c.number_format = "#,##0"
        row += 1

    # Totals
    _write_totals(ws, row, summary)


def _write_totals(ws, start_row: int, summary: EstimationSummary):
    ws.cell(row=start_row + 1, column=7, value="小計").font = Font(bold=True)
    ws.cell(row=start_row + 1, column=8, value=float(summary.subtotal)).number_format = "#,##0"
    ws.cell(row=start_row + 2, column=7, value="諸経費").font = Font(bold=True)
    ws.cell(row=start_row + 2, column=8, value=float(summary.misc_fee)).number_format = "#,##0"
    ws.cell(row=start_row + 3, column=7, value="消費税").font = Font(bold=True)
    ws.cell(row=start_row + 3, column=8, value=float(summary.tax)).number_format = "#,##0"
    ws.cell(row=start_row + 4, column=7, value="合計（税込）").font = Font(bold=True, size=12)
    ws.cell(row=start_row + 4, column=8, value=float(summary.grand_total)).number_format = "#,##0"
    ws.cell(row=start_row + 4, column=8).font = Font(bold=True, size=12)

    # Stamp area
    ws.merge_cells(f"A{start_row + 2}:C{start_row + 5}")
    ws[f"A{start_row + 2}"] = "承認　　　　　　　担当"
    ws[f"A{start_row + 2}"].alignment = Alignment(horizontal="center", vertical="center")


def _write_order_sheet(ws, summary: EstimationSummary):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14

    headers = ["メーカー", "型番", "品名", "数量", "単位", "参考単価"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = Font(color="FFFFFF", bold=True)
        c.border = _border

    items = sorted(
        [i for i in summary.confirmed_items if i.maker],
        key=lambda x: (x.maker or "", x.model_number or "")
    )
    for row, item in enumerate(items, 2):
        for col, val in enumerate([item.maker, item.model_number, item.item_name,
                                    float(item.quantity), item.unit, float(item.unit_price)], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _border
            if col == 6:
                c.number_format = "#,##0"


def _write_pending_sheet(ws, summary: EstimationSummary):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40

    headers = ["記号コード", "品名", "備考・確認事項"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="FF0000")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = _border

    for row, item in enumerate(summary.pending_items, 2):
        for col, val in enumerate([item.symbol_code, item.item_name, item.note or ""], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PENDING_FILL
            c.border = _border


def build_pdf(summary: EstimationSummary, company_name: str, site_name: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, spaceAfter=6)
    normal = styles["Normal"]

    story = []
    story.append(Paragraph("見　積　書", title_style))
    story.append(Paragraph(f"会社名: {company_name}　　現場名: {site_name}", normal))
    story.append(Paragraph(
        f"作成日: {date.today().strftime('%Y年%m月%d日')}　　"
        f"有効期限: {(date.today() + timedelta(days=30)).strftime('%Y年%m月%d日')}",
        normal))
    story.append(Spacer(1, 6*mm))

    # Table data
    col_headers = ["記号", "品名", "数量", "単位", "単価", "金額"]
    data = [col_headers]
    all_items = summary.confirmed_items + summary.pending_items
    for item in all_items:
        label = item.item_name + (" [要確認]" if item.status == "PENDING" else "")
        data.append([
            item.symbol_code,
            label,
            f"{float(item.quantity):.1f}",
            item.unit,
            f"{float(item.unit_price):,.0f}",
            f"{float(item.amount):,.0f}",
        ])

    # Totals rows
    data.append(["", "", "", "", "小計", f"{float(summary.subtotal):,.0f}"])
    data.append(["", "", "", "", "諸経費", f"{float(summary.misc_fee):,.0f}"])
    data.append(["", "", "", "", "消費税", f"{float(summary.tax):,.0f}"])
    data.append(["", "", "", "", "合計（税込）", f"{float(summary.grand_total):,.0f}"])

    col_widths = [25*mm, 70*mm, 18*mm, 15*mm, 22*mm, 25*mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -len(data)+len(summary.confirmed_items)+1),
         [colors.white, colors.HexColor("#F2F2F2")]),
        ("BACKGROUND", (0, -4), (-1, -1), colors.HexColor("#EBF3FB")),
        ("FONTNAME", (4, -1), (5, -1), "Helvetica-Bold"),
    ]))
    story.append(t)

    doc.build(story)
    return buf.getvalue()
